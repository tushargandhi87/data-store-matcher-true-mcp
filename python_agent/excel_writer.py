"""Excel Writer for generating output files."""
import logging
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Write results to Excel files with formatting."""
    
    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
    def _format_worksheet(self, filepath: Path):
        """Apply formatting to Excel worksheet."""
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
            logger.info(f"Formatted Excel file: {filepath}")
            
        except Exception as e:
            logger.error(f"Error formatting Excel file: {e}", exc_info=True)
    
    def write_match_results(self, results: List[Dict[str, Any]], filename: str = "datastore_match_results.xlsx"):
        """Write datastore matching results to Excel."""
        try:
            logger.info(f"Writing {len(results)} match results to {filename}")
            
            rows = []
            for result in results:
                row = {
                    "Input Datastore": result.get("input_datastore", ""),
                    "Matched Datastore": result.get("matched_datastore", ""),
                    "Confidence Score": result.get("confidence", 0.0),
                    "Reasoning": result.get("reasoning", ""),
                    "Requires EOL Lookup": "Yes" if result.get("confidence", 1.0) < 0.7 else "No",
                    "Processing Status": "Completed",
                    "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                rows.append(row)
            
            df = pd.DataFrame(rows)
            filepath = self.output_dir / filename
            df.to_excel(filepath, index=False, sheet_name="Match Results")
            
            self._format_worksheet(filepath)
            
            logger.info(f"Successfully wrote match results to {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error writing match results: {e}", exc_info=True)
            raise
    
    def write_eol_success(self, results: List[Dict[str, Any]], filename: str = "api_success.xlsx"):
        """Write successful EOL lookup results to Excel."""
        try:
            logger.info(f"Writing {len(results)} EOL success results to {filename}")
            
            rows = []
            for result in results:
                row = {
                    "Input Datastore": result.get("input_datastore", ""),
                    "Product": result.get("product", ""),
                    "Version": result.get("version", ""),
                    "API Product Name": result.get("api_product_name", ""),
                    "API Matched Version": result.get("matched_version", ""),
                    "Match Type": result.get("match_type", ""),
                    "EOL Date": result.get("eol_date", ""),
                    "Support Status": result.get("support_status", ""),
                    "Latest Version": result.get("latest_version", ""),
                    "LTS Version": result.get("lts_version", ""),
                    "Release Date": result.get("release_date", "")
                }
                rows.append(row)
            
            df = pd.DataFrame(rows)
            filepath = self.output_dir / filename
            df.to_excel(filepath, index=False, sheet_name="API Success")
            
            self._format_worksheet(filepath)
            
            logger.info(f"Successfully wrote EOL success results to {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error writing EOL success results: {e}", exc_info=True)
            raise
    
    def write_eol_not_found(self, results: List[Dict[str, Any]], filename: str = "api_not_found.xlsx"):
        """Write EOL lookup not found results to Excel."""
        try:
            logger.info(f"Writing {len(results)} EOL not found results to {filename}")
            
            rows = []
            for result in results:
                available_versions = result.get("available_versions", [])
                if isinstance(available_versions, list):
                    available_versions = ", ".join(available_versions)
                
                row = {
                    "Input Datastore": result.get("input_datastore", ""),
                    "Product": result.get("product", ""),
                    "Version": result.get("version", ""),
                    "API Product Name": result.get("api_product_name", ""),
                    "Not Found Type": result.get("error_type", ""),
                    "Available Versions": available_versions,
                    "Error Message": result.get("error_message", "")
                }
                rows.append(row)
            
            df = pd.DataFrame(rows)
            filepath = self.output_dir / filename
            df.to_excel(filepath, index=False, sheet_name="API Not Found")
            
            self._format_worksheet(filepath)
            
            logger.info(f"Successfully wrote EOL not found results to {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error writing EOL not found results: {e}", exc_info=True)
            raise
    
    def write_eol_errors(self, results: List[Dict[str, Any]], filename: str = "api_errors.xlsx"):
        """Write EOL lookup error results to Excel."""
        try:
            logger.info(f"Writing {len(results)} EOL error results to {filename}")
            
            rows = []
            for result in results:
                row = {
                    "Input Datastore": result.get("input_datastore", ""),
                    "Product": result.get("product", ""),
                    "Version": result.get("version", ""),
                    "API Product Name": result.get("api_product_name", ""),
                    "Error Type": result.get("error_type", ""),
                    "Error Details": result.get("error_message", ""),
                    "Retry Count": result.get("retry_count", 0),
                    "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                rows.append(row)
            
            df = pd.DataFrame(rows)
            filepath = self.output_dir / filename
            df.to_excel(filepath, index=False, sheet_name="API Errors")
            
            self._format_worksheet(filepath)
            
            logger.info(f"Successfully wrote EOL error results to {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error writing EOL error results: {e}", exc_info=True)
            raise
